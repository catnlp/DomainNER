# encoding:utf-8
'''
@Author: catnlp
@Email: wk_nlp@163.com
@Time: 2018/5/24 15:48
'''
import os
import json
import openpyxl

class Origin2Raw:
    def __init__(self, path, save):
        if os.path.exists(path):
            if path[-1] != '/':
                path += '/'
            self.path = path
        else:
            print('Folder does not exists!')
            exit(1)

        if not os.path.exists(save):
            os.makedirs(save)
        if save[-1] != '/':
            save += '/'
        self.save = save

        print('Origin path: %s', self.path)
        print('Raw path: %s', self.save)

    def excel2dict(self, dict_name, save_name='dict.txt', sheet_name=None):
        dict_path = self.path + dict_name
        print('Excel path: %s', dict_path)
        if not os.path.exists(dict_path):
            print('Excel does not exists!')
            exit(1)
        save_path = self.save + save_name

        work_book = openpyxl.load_workbook(dict_path)
        if sheet_name:
            sheet = work_book.get_sheet_by_name(sheet_name)
        else:
            sheet_names = work_book.get_sheet_names()
            sheet = work_book.get_sheet_by_name(sheet_names[0])

        with open(save_path, 'w') as dict:
            for row in range(1, sheet.max_row):
                dict.write(sheet.cell(row+1, 2).value + '\t' + sheet.cell(row+1, 3).value + '\n')
                print(sheet.cell(row+1, 2).value + '\t' + sheet.cell(row+1, 3).value)

    def json2text(self, save_name='raw.txt', json_suffix='jsonl'):
        save_path = self.save + save_name
        with open(save_path, 'w') as raw:
            for parent, dirnames, filenames in os.walk(self.path):
                for filename in filenames:
                    if filename.endswith(json_suffix):
                        file_path = os.path.join(parent, filename)
                        print('文件名：%s' % filename)
                        print('文件完整路径：%s\n' % file_path)
                        with open(file_path) as origin:
                            lines = origin.readlines()
                            for line in lines:
                                dict = json.loads(line)
                                if dict['answer'] == 'accept':
                                    raw.write(dict['text'] + '\n')

if __name__ == '__main__':
    origin2raw = Origin2Raw('../data/origin/Anti-fraud Product Data', '../data/raw')
    origin2raw.excel2dict('Entity Dictionary.xlsx')
    origin2raw.json2text()